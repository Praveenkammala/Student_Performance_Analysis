import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference

# Load cleaned data
students = pd.read_csv("../data/cleaned/students_cleaned.csv")
marks = pd.read_csv("../data/cleaned/marks_cleaned.csv")
exams = pd.read_csv("../data/cleaned/exams_cleaned.csv")

wb = Workbook()

# ---------------- Students Sheet ----------------
ws1 = wb.active
ws1.title = "Students_Data"
ws1.append(list(students.columns))
for row in students.itertuples(index=False):
    ws1.append(list(row))

# ---------------- Marks Sheet ----------------
ws2 = wb.create_sheet("Marks_Data")
ws2.append(list(marks.columns))
for row in marks.itertuples(index=False):
    ws2.append(list(row))

# ---------------- Subject Analysis ----------------
ws3 = wb.create_sheet("Subject_Analysis")
avg_subject = marks.groupby("subject")["marks"].mean().reset_index()
ws3.append(["Subject", "Average Marks"])
for row in avg_subject.itertuples(index=False):
    ws3.append(list(row))

bar1 = BarChart()
bar1.title = "Average Marks by Subject"
data = Reference(ws3, min_col=2, min_row=1, max_row=len(avg_subject)+1)
cats = Reference(ws3, min_col=1, min_row=2, max_row=len(avg_subject)+1)
bar1.add_data(data, titles_from_data=True)
bar1.set_categories(cats)
ws3.add_chart(bar1, "E2")

# ---------------- Department Analysis ----------------
ws4 = wb.create_sheet("Department_Analysis")
merged = marks.merge(students, on="student_id")
avg_dept = merged.groupby("department")["marks"].mean().reset_index()
ws4.append(["Department", "Average Marks"])
for row in avg_dept.itertuples(index=False):
    ws4.append(list(row))

bar2 = BarChart()
bar2.title = "Average Marks by Department"
data = Reference(ws4, min_col=2, min_row=1, max_row=len(avg_dept)+1)
cats = Reference(ws4, min_col=1, min_row=2, max_row=len(avg_dept)+1)
bar2.add_data(data, titles_from_data=True)
bar2.set_categories(cats)
ws4.add_chart(bar2, "E2")

# ---------------- Attendance Impact ----------------
ws5 = wb.create_sheet("Attendance_Impact")
ws5.append(["Attendance", "Marks"])
for row in marks[["attendance", "marks"]].itertuples(index=False):
    ws5.append(list(row))

scatter = ScatterChart()
scatter.title = "Attendance vs Marks"
xvalues = Reference(ws5, min_col=1, min_row=2, max_row=len(marks)+1)
yvalues = Reference(ws5, min_col=2, min_row=2, max_row=len(marks)+1)
series = scatter.series.append
scatter.series.append(yvalues)
scatter.x_axis.title = "Attendance (%)"
scatter.y_axis.title = "Marks"
ws5.add_chart(scatter, "E2")

# ---------------- Pass vs Fail ----------------
ws6 = wb.create_sheet("Results")
result_counts = exams["result"].value_counts().reset_index()
ws6.append(["Result", "Count"])
for row in result_counts.itertuples(index=False):
    ws6.append(list(row))

pie = PieChart()
pie.title = "Pass vs Fail Distribution"
data = Reference(ws6, min_col=2, min_row=1, max_row=3)
labels = Reference(ws6, min_col=1, min_row=2, max_row=3)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
ws6.add_chart(pie, "E2")

# Save
wb.save("../excel/Student_Performance_Report.xlsx")

print("Advanced Excel dashboard generated successfully")
